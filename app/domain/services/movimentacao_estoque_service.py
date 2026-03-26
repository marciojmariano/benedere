"""
Service: MovimentacaoEstoque
Responsabilidade: regras de negócio e casos de uso do estoque.
"""
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl

from app.api.v1.schemas.movimentacao_estoque import ImportacaoEstoqueResponse, ImportacaoLinhaErro
from app.infra.database.models.base import TipoMovimentacao
from app.infra.database.models.movimentacao_estoque import MovimentacaoEstoque
from app.infra.repository.movimentacao_estoque_repository import MovimentacaoEstoqueRepository
from app.infra.repository.ingrediente_repository import IngredienteRepository


# ── Exceções de domínio ───────────────────────────────────────────────────────

class MovimentacaoNaoEncontradaError(Exception):
    def __init__(self, mov_id: uuid.UUID):
        super().__init__(f"Movimentação não encontrada: {mov_id}")


class IngredienteNaoEncontradoParaEstoqueError(Exception):
    def __init__(self, ingrediente_id: uuid.UUID):
        super().__init__(f"Ingrediente não encontrado ou inativo: {ingrediente_id}")


# ── Service ───────────────────────────────────────────────────────────────────

class MovimentacaoEstoqueService:

    def __init__(
        self,
        mov_repo: MovimentacaoEstoqueRepository,
        ingrediente_repo: IngredienteRepository,
        tenant_id: uuid.UUID,
    ) -> None:
        self._mov_repo = mov_repo
        self._ingrediente_repo = ingrediente_repo
        self._tenant_id = tenant_id

    async def registrar_entrada(
        self,
        ingrediente_id: uuid.UUID,
        quantidade: Decimal,
        preco_unitario_custo: Decimal,
        data_movimentacao: date,
        observacoes: str | None = None,
    ) -> MovimentacaoEstoque:
        # Valida ingrediente
        ingrediente = await self._ingrediente_repo.get_by_id(ingrediente_id)
        if not ingrediente or not ingrediente.ativo:
            raise IngredienteNaoEncontradoParaEstoqueError(ingrediente_id)

        # Cria movimentação
        mov = MovimentacaoEstoque(
            tenant_id=self._tenant_id,
            ingrediente_id=ingrediente_id,
            tipo=TipoMovimentacao.COMPRA,
            quantidade=quantidade,
            preco_unitario_custo=preco_unitario_custo,
            data_movimentacao=data_movimentacao,
            observacoes=observacoes,
        )

        # Atualiza saldo e custo do ingrediente (atomicidade garantida pelo flush único)
        ingrediente.saldo_atual = Decimal(str(ingrediente.saldo_atual)) + quantidade
        ingrediente.custo_unitario = float(preco_unitario_custo)
        ingrediente.updated_at = datetime.utcnow()

        return await self._mov_repo.create(mov)

    async def buscar_por_id(self, mov_id: uuid.UUID) -> MovimentacaoEstoque:
        mov = await self._mov_repo.get_by_id(mov_id)
        if not mov:
            raise MovimentacaoNaoEncontradaError(mov_id)
        return mov

    async def listar_todas(self, limit: int = 50, offset: int = 0) -> list[MovimentacaoEstoque]:
        return await self._mov_repo.list_all(limit=limit, offset=offset)

    async def listar_por_ingrediente(
        self,
        ingrediente_id: uuid.UUID,
        limit: int = 50,
        offset: int = 0,
    ) -> list[MovimentacaoEstoque]:
        return await self._mov_repo.list_by_ingrediente(
            ingrediente_id=ingrediente_id, limit=limit, offset=offset
        )

    async def importar_excel(self, conteudo: bytes) -> ImportacaoEstoqueResponse:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active

        # Lê cabeçalho e mapeia nome da coluna → índice (strip para remover \xa0 e espaços)
        cabecalho = {
            str(cell.value).strip().lower(): idx
            for idx, cell in enumerate(ws[1])
            if cell.value is not None
        }
        COLUNAS_OBRIGATORIAS = {"ingrediente_nome", "quantidade", "preco_unitario_custo", "data_movimentacao"}
        ausentes = COLUNAS_OBRIGATORIAS - cabecalho.keys()
        if ausentes:
            return ImportacaoEstoqueResponse(
                total_linhas=0,
                importadas=0,
                erros=[ImportacaoLinhaErro(
                    linha=1,
                    ingrediente_nome="",
                    mensagem=f"Colunas obrigatórias ausentes no cabeçalho: {', '.join(sorted(ausentes))}",
                )],
            )

        def col(row_values: tuple, nome: str):
            idx = cabecalho.get(nome)
            return row_values[idx] if idx is not None and idx < len(row_values) else None

        importadas = 0
        erros: list[ImportacaoLinhaErro] = []

        # Linha 1 é o cabeçalho; dados começam na linha 2
        for num_linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignora linhas completamente vazias
            if not any(row):
                continue

            ingrediente_nome_raw = str(col(row, "ingrediente_nome")).strip() if col(row, "ingrediente_nome") is not None else ""
            quantidade_raw = col(row, "quantidade")
            preco_raw = col(row, "preco_unitario_custo")
            data_raw = col(row, "data_movimentacao")
            obs_val = col(row, "observacoes")
            observacoes_raw = str(obs_val).strip() if obs_val is not None else None

            # Valida campos obrigatórios
            if not ingrediente_nome_raw:
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha, ingrediente_nome="", mensagem="Nome do ingrediente obrigatório"
                ))
                continue

            # Busca ingrediente por nome
            ingrediente = await self._ingrediente_repo.get_by_nome(ingrediente_nome_raw)
            if not ingrediente:
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha,
                    ingrediente_nome=ingrediente_nome_raw,
                    mensagem=f"Ingrediente '{ingrediente_nome_raw}' não encontrado ou inativo",
                ))
                continue

            # Valida e converte quantidade
            try:
                quantidade = round(Decimal(str(quantidade_raw)), 4)
                if quantidade <= 0:
                    raise ValueError()
            except (InvalidOperation, ValueError, TypeError):
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha,
                    ingrediente_nome=ingrediente_nome_raw,
                    mensagem=f"Quantidade inválida: '{quantidade_raw}'",
                ))
                continue

            # Valida e converte preço
            try:
                preco = round(Decimal(str(preco_raw)), 4)
                if preco <= 0:
                    raise ValueError()
            except (InvalidOperation, ValueError, TypeError):
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha,
                    ingrediente_nome=ingrediente_nome_raw,
                    mensagem=f"Preço unitário inválido: '{preco_raw}'",
                ))
                continue

            # Valida e converte data
            try:
                if isinstance(data_raw, datetime):
                    data_mov = data_raw.date()
                elif isinstance(data_raw, date):
                    data_mov = data_raw
                elif isinstance(data_raw, str):
                    data_str = data_raw.strip()
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                        try:
                            data_mov = datetime.strptime(data_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        raise ValueError(f"formato inválido: '{data_str}'")
                else:
                    raise ValueError(f"tipo inesperado: {type(data_raw)}")
            except (ValueError, TypeError) as exc:
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha,
                    ingrediente_nome=ingrediente_nome_raw,
                    mensagem=f"Data inválida: {exc}",
                ))
                continue

            # Registra entrada reutilizando a lógica existente
            try:
                await self.registrar_entrada(
                    ingrediente_id=ingrediente.id,
                    quantidade=quantidade,
                    preco_unitario_custo=preco,
                    data_movimentacao=data_mov,
                    observacoes=observacoes_raw or None,
                )
                importadas += 1
            except Exception as exc:
                erros.append(ImportacaoLinhaErro(
                    linha=num_linha,
                    ingrediente_nome=ingrediente_nome_raw,
                    mensagem=str(exc),
                ))

        return ImportacaoEstoqueResponse(
            total_linhas=importadas + len(erros),
            importadas=importadas,
            erros=erros,
        )
